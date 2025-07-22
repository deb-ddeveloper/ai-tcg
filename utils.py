import pytesseract
from PIL import Image
import openpyxl
from openpyxl import Workbook

def extract_text_from_image(image_path):
    try:
        text = pytesseract.image_to_string(Image.open(image_path))
        return text
    except Exception as e:
        raise Exception(f"Error extracting text from image: {str(e)}")

def generate_testcases_from_text(text, count):
    try:
        lines = text.splitlines()
        test_cases = []
        for i, line in enumerate(lines):
            if len(test_cases) >= count:
                break
            if line.strip():
                test_cases.append({"title": f"Test {i+1}", "steps": line.strip(), "expected": "Expected outcome"})
        return test_cases
    except Exception as e:
        raise Exception(f"Error generating test cases from text: {str(e)}")

def parse_template(template_path):
    try:
        wb = openpyxl.load_workbook(template_path)
        sheet = wb.active
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1)) if cell.value]
        return headers
    except Exception as e:
        raise Exception(f"Failed to parse template: {str(e)}")

def generate_testcase_file(cases, columns, output_path):
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(columns)
        for c in cases:
            row = [c.get(col, '') for col in columns]
            ws.append(row)
        wb.save(output_path)
    except Exception as e:
        raise Exception(f"Failed to generate test case file: {str(e)}")